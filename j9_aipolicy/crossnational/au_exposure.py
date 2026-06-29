"""AU exposure: international-student share per institution = (all - domestic)/all,
from AU Dept of Education 2022 Section 2 tables 2.5 (All) and 2.6 (All Domestic),
TOTAL column. Match to the AU institutions in our sample and fill intl_share."""
import csv, re
import openpyxl, os
HERE = os.path.dirname(os.path.abspath(__file__))

P = os.path.join(HERE, "2022 Section 2 - All students.xlsx")
FRAME = os.path.join(HERE, "data", "sample_frame.csv")


def norm(s):
    s = re.sub(r"\bthe\b", "", str(s).lower())
    s = re.sub(r"[^a-z]+", " ", s).strip()
    return re.sub(r"\s+", " ", s)


def totals(sheetname):
    wb = openpyxl.load_workbook(P, read_only=True, data_only=True)
    ws = wb[sheetname]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    header = [str(c) if c is not None else "" for c in rows[2]]
    inst_col = 1
    tot_col = max(i for i, h in enumerate(header) if h.strip().upper() == "TOTAL")
    out = {}
    for r in rows[3:]:
        inst = r[inst_col]
        if not inst:
            continue
        try:
            tot = float(r[tot_col])
        except (TypeError, ValueError):
            continue
        out[norm(inst)] = tot
    return out


allt = totals("2.5")
domt = totals("2.6")
share = {k: (allt[k] - domt[k]) / allt[k] for k in allt if k in domt and allt[k] > 0}

frame = list(csv.DictReader(open(FRAME)))
cols = list(frame[0].keys())
matched, unmatched = [], []
for r in frame:
    if r["country"] != "AU":
        continue
    key = norm(r["institution"])
    if key in share:
        r["intl_share"] = round(share[key], 4)
        r["exposure_year"] = 2022
        matched.append((r["institution"], r["intl_share"]))
    else:
        unmatched.append(r["institution"])

with open(FRAME, "w", newline="") as f:
    w = csv.DictWriter(f, fieldnames=cols); w.writeheader(); w.writerows(frame)

print(f"AU institutions in sample matched to gov data: {len(matched)}")
for n, s in sorted(matched, key=lambda x: -x[1]):
    print(f"   {n:<34} intl_share = {s:.1%}")
if unmatched:
    print("UNMATCHED:", unmatched)
